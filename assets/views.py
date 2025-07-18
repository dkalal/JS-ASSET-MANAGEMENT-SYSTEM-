from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, DetailView, TemplateView, UpdateView
from .models import Asset, AssetCategory, AssetCategoryField, ExportLog
from .forms import AssetForm
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import re
from django.http import HttpResponse
import csv
import pandas as pd
from django.conf import settings
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile
from datetime import datetime
import openpyxl
from django.core.files.storage import default_storage
from django.contrib import messages
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from audit.models import AuditLog
from audit.utils import log_audit, ASSIGN_ACTION, MAINTENANCE_ACTION
from django.core.paginator import Paginator, EmptyPage
from django.utils.timezone import localtime

# Permission check: only admin/manager
def is_admin_or_manager(user):
    return user.is_authenticated and user.role in ('admin', 'manager')

# Create your views here.

class AssetCreateView(UserPassesTestMixin, CreateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('asset_list')

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role in ('admin', 'manager')

    def form_valid(self, form):
        asset = form.save(commit=False)
        assigned_to = form.cleaned_data.get('assigned_to')
        asset.save()  # Save first to ensure UUID is set
        # Generate QR code with direct URL
        base_url = self.request.build_absolute_uri('/')[:-1]  # Remove trailing slash
        qr_url = f"{base_url}/assets/{asset.uuid}/"
        qr = qrcode.make(qr_url)
        buffer = BytesIO()
        qr.save(buffer, 'PNG')
        asset.qr_code.save(f"asset_{asset.uuid}.png", ContentFile(buffer.getvalue()), save=False)
        asset.save()
        if assigned_to:
            log_audit(self.request.user, ASSIGN_ACTION, asset, f'Asset assigned to {assigned_to}', related_user=assigned_to)
        log_audit(self.request.user, 'create', asset, 'Asset created via dashboard')
        messages.success(self.request, f"Asset '{asset}' registered successfully.")
        print(f"[DEBUG] Asset created: {asset}")
        return super().form_valid(form)

    def form_invalid(self, form):
        print(f"[DEBUG] Asset registration form invalid: {form.errors}")
        messages.error(self.request, "Asset registration failed. Please correct the errors below.")
        return super().form_invalid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial'] = kwargs.get('initial', {})
        kwargs['initial']['request'] = self.request
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_role'] = getattr(self.request.user, 'role', 'user')
        return context

asset_create = user_passes_test(is_admin_or_manager, login_url='login')(AssetCreateView.as_view())

# Asset update view for admin/manager with audit logging
class AssetUpdateView(UserPassesTestMixin, UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('asset_list')
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role in ('admin', 'manager')

    def form_valid(self, form):
        old_obj = self.get_object()
        asset = form.save(commit=False)
        assigned_to = form.cleaned_data.get('assigned_to')
        status = form.cleaned_data.get('status')
        asset.save()
        # Assignment logging: if assigned_to changes
        if old_obj.assigned_to != assigned_to and assigned_to:
            log_audit(self.request.user, ASSIGN_ACTION, asset, f'Asset assigned to {assigned_to}', related_user=assigned_to)
        # Transfer logging: if status changes to 'transferred'
        if old_obj.status != status and status == 'transferred':
            log_audit(
                self.request.user,
                ASSIGN_ACTION,
                asset,
                f"Asset status set to 'transferred' (from {old_obj.assigned_to} to {assigned_to})",
                related_user=assigned_to
            )
        # Maintenance logging
        if old_obj.status != status and status == 'maintenance':
            log_audit(self.request.user, MAINTENANCE_ACTION, asset, 'Asset marked as under maintenance')
        return super().form_valid(form)

@require_GET
def get_dynamic_fields(request):
    category_id = request.GET.get('category_id')
    try:
        category = AssetCategory.objects.get(pk=category_id)
        return JsonResponse({'success': True, 'fields': category.dynamic_fields})
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'fields': {}})

# Asset list view: only for authenticated users
class AssetListView(LoginRequiredMixin, ListView):
    model = Asset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('category', 'assigned_to')
        user = self.request.user
        role = getattr(user, 'role', 'user')
        # Enforce role-based filtering
        if role == 'user':
            qs = qs.filter(assigned_to=user)
        category = self.request.GET.get('category')
        status = self.request.GET.get('status')
        location = self.request.GET.get('location')
        search = self.request.GET.get('search')
        assigned = self.request.GET.get('assigned')
        warranty = self.request.GET.get('warranty')
        # Dynamic field filters
        dynamic_filters = {}
        if category:
            qs = qs.filter(category__id=category)
            # Fetch dynamic fields for this category
            fields = AssetCategoryField.objects.filter(category_id=category)
            for field in fields:
                val = self.request.GET.get(f'dyn_{field.key}')
                if val:
                    # Filter by dynamic_data JSON key
                    if field.type == 'text':
                        qs = qs.filter(**{f'dynamic_data__{field.key}__icontains': val})
                    elif field.type == 'number':
                        try:
                            qs = qs.filter(**{f'dynamic_data__{field.key}': float(val)})
                        except ValueError:
                            pass
                    elif field.type == 'date':
                        # Parse mm/dd/yyyy and convert to yyyy-MM-dd
                        import datetime
                        try:
                            dt = datetime.datetime.strptime(val, '%m/%d/%Y')
                            iso_val = dt.strftime('%Y-%m-%d')
                            qs = qs.filter(**{f'dynamic_data__{field.key}': iso_val})
                        except ValueError:
                            pass  # Invalid date format, ignore filter
        if status:
            qs = qs.filter(status=status)
        if assigned == 'yes':
            qs = qs.filter(assigned_to__isnull=False)
        elif assigned == 'no':
            qs = qs.filter(assigned_to__isnull=True)
        if warranty == 'expiring':
            from datetime import timedelta
            from django.utils import timezone
            soon = timezone.now() + timedelta(days=30)
            qs = qs.filter(dynamic_data__warranty_expiry__lte=soon.date().isoformat(), dynamic_data__warranty_expiry__gte=timezone.now().date().isoformat())
        if location:
            qs = qs.filter(dynamic_data__location__icontains=location)
        if search:
            qs = qs.filter(
                Q(dynamic_data__name__icontains=search) |
                Q(dynamic_data__model__icontains=search) |
                Q(description__icontains=search)
            )
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = AssetCategory.objects.all()
        context['statuses'] = Asset.STATUS_CHOICES
        selected_category = self.request.GET.get('category')
        if selected_category:
            # Show all dynamic fields for the selected category
            context['dynamic_fields'] = AssetCategoryField.objects.filter(category_id=selected_category)
        else:
            # Show the union of all dynamic fields across all categories, deduplicated by key and label (case-insensitive)
            all_fields = AssetCategoryField.objects.all()
            seen = set()
            unique_fields = []
            for f in all_fields:
                dedup_key = (f.key.lower().strip(), f.label.lower().strip())
                if dedup_key not in seen:
                    unique_fields.append(f)
                    seen.add(dedup_key)
            context['dynamic_fields'] = unique_fields
        return context

class AssetDetailView(LoginRequiredMixin, DetailView):
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'

    def get(self, request, *args, **kwargs):
        asset = self.get_object()
        log_audit(request.user, 'view', asset, 'Asset viewed via dashboard')
        # Redirect to UUID-based URL if accessed by PK
        return redirect('asset_detail_by_uuid', uuid=asset.uuid)

class AssetScanView(TemplateView):
    template_name = 'assets/asset_scan.html'

@require_GET
@csrf_exempt
def asset_by_code(request):
    code = request.GET.get('code')
    asset = None
    if code:
        # Try to extract UUID from QR code format 'ASSET|v1|{uuid}'
        match = re.match(r'^ASSET\|v1\|([0-9a-fA-F-]{36})$', code)
        if match:
            uuid_val = match.group(1)
            asset = Asset.objects.filter(uuid=uuid_val).first()
        # Fallback: Try by QR code filename or asset ID
        if not asset:
            asset = Asset.objects.filter(qr_code__icontains=code).first()
        if not asset and code.isdigit():
            asset = Asset.objects.filter(pk=int(code)).first()
    if asset:
        # Log QR code scan
        log_audit(request.user, 'scan', asset, f'QR code scanned: {code}')
        data = {
            'id': asset.pk,
            'dynamic_data': asset.dynamic_data,
            'category_name': asset.category.name,
            'status': asset.status,
            'assigned_to': str(asset.assigned_to) if asset.assigned_to else '',
            'created_at': asset.created_at.strftime('%Y-%m-%d %H:%M'),
        }
        return JsonResponse({'success': True, 'asset': data})
    return JsonResponse({'success': False})

class AssetDetailByUUIDView(LoginRequiredMixin, DetailView):
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['accessed_by_uuid'] = True
        return context

# Export endpoint (robust, supports GET and POST, individual/bulk)
def asset_export(request):
    if request.method == 'POST':
        format = request.POST.get('format', 'csv')
        columns = request.POST.getlist('columns')
        selected_ids = request.POST.get('selected_ids', '')
    else:
        format = request.GET.get('format', 'csv')
        columns = request.GET.getlist('columns')
        selected_ids = request.GET.get('selected_ids', '')
    # Reuse AssetListView filtering logic
    assets = Asset.objects.all()
    # If selected_ids provided, filter by those
    if selected_ids:
        id_list = [int(pk) for pk in selected_ids.split(',') if pk.strip().isdigit()]
        assets = assets.filter(pk__in=id_list)
    else:
        category = request.GET.get('category')
        status = request.GET.get('status')
        location = request.GET.get('location')
        search = request.GET.get('search')
        if category:
            assets = assets.filter(category__id=category)
            fields = AssetCategoryField.objects.filter(category_id=category)
            for field in fields:
                val = request.GET.get(f'dyn_{field.key}')
                if val:
                    if field.type == 'text':
                        assets = assets.filter(**{f'dynamic_data__{field.key}__icontains': val})
                    elif field.type == 'number':
                        try:
                            assets = assets.filter(**{f'dynamic_data__{field.key}': float(val)})
                        except ValueError:
                            pass
                    elif field.type == 'date':
                        assets = assets.filter(**{f'dynamic_data__{field.key}': val})
        if status:
            assets = assets.filter(status=status)
        if location:
            assets = assets.filter(dynamic_data__location__icontains=location)
        if search:
            assets = assets.filter(
                Q(dynamic_data__name__icontains=search) |
                Q(dynamic_data__model__icontains=search) |
                Q(description__icontains=search)
            )
    # Prepare data
    data = []
    for asset in assets:
        row = {}
        # Core fields
        row['ID'] = asset.pk
        row['Category'] = asset.category.name
        row['Status'] = asset.status
        row['Assigned To'] = str(asset.assigned_to) if asset.assigned_to else ''
        row['Created'] = asset.created_at.strftime('%Y-%m-%d %H:%M')
        row['Updated'] = asset.updated_at.strftime('%Y-%m-%d %H:%M')
        # Dynamic fields
        for k, v in asset.dynamic_data.items():
            row[k] = v
        data.append(row)
    # Filter columns
    if columns:
        data = [ {k: row.get(k, '') for k in columns} for row in data ]
    # Large export warning
    large_export = len(data) > 1000
    # Log export
    log = ExportLog.objects.create(
        user=request.user if request.user.is_authenticated else None,
        format=format,
        columns=columns,
        filters=request.POST.dict() if request.method == 'POST' else request.GET.dict(),
        success=True
    )
    try:
        if format == 'csv':
            df = pd.DataFrame(data)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"'
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            df.to_csv(response, index=False)
            log_audit(request.user, 'export', None, 'Assets exported as CSV')
            return response
        elif format == 'xlsx':
            df = pd.DataFrame(data)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"'
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Assets')
            log_audit(request.user, 'export', None, 'Assets exported as Excel')
            return response
        elif format == 'pdf':
            html_string = render_to_string('assets/export_pdf.html', {
                'assets': data,
                'columns': columns or (data[0].keys() if data else []),
                'logo_url': settings.STATIC_URL + 'img/logo.png',
                'export_date': datetime.now(),
            })
            import os
            fd, temp_path = tempfile.mkstemp(suffix='.pdf')
            os.close(fd)
            try:
                HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(temp_path)
                with open(temp_path, 'rb') as f:
                    pdf = f.read()
            finally:
                os.remove(temp_path)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="assets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"'
            if large_export:
                response['X-Export-Warning'] = 'Export is very large and may take time.'
            log_audit(request.user, 'export', None, 'Assets exported as PDF')
            return response
        else:
            log.success = False
            log.error_message = 'Invalid export format'
            log.save()
            return HttpResponse('Invalid export format', status=400)
    except Exception as e:
        log.success = False
        log.error_message = str(e)
        log.save()
        return HttpResponse(f'Export failed: {e}', status=500)

@login_required
@user_passes_test(is_admin_or_manager, login_url='login')
def download_import_template(request):
    # Generate Excel template for selected category
    category_id = request.GET.get('category')
    if not category_id:
        return HttpResponse('Category required', status=400)
    category = AssetCategory.objects.get(pk=category_id)
    dynamic_fields = AssetCategoryField.objects.filter(category=category)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Core fields
    columns = ['status', 'description', 'assigned_to']
    # Dynamic fields
    columns += [f.key for f in dynamic_fields]
    ws.append(columns)
    # Add sample row
    if request.GET.get('example') == '1':
        # Provide realistic example data
        sample = ['active', 'Laptop for CEO', 'manager']
        for f in dynamic_fields:
            if f.key == 'serial_number':
                sample.append('SN123456')
            elif f.key == 'location':
                sample.append('HQ Office')
            elif f.key == 'purchase_date':
                sample.append('2023-01-15')
            else:
                sample.append('Example')
        ws.append(sample)
    else:
        sample = ['active', 'Sample asset', 'manager'] + ['' for _ in dynamic_fields]
        ws.append(sample)
    # Style header
    for i, col in enumerate(columns, 1):
        ws[f'{get_column_letter(i)}1'].font = openpyxl.styles.Font(bold=True)
    # Return as response
    import os
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    try:
        wb.save(temp_path)
        with open(temp_path, 'rb') as f:
            file_data = f.read()
    finally:
        os.remove(temp_path)
    response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if request.GET.get('example') == '1':
        response['Content-Disposition'] = f'attachment; filename=asset_import_example_{category.name}.xlsx'
    else:
        response['Content-Disposition'] = f'attachment; filename=asset_import_template_{category.name}.xlsx'
    return response

@method_decorator(user_passes_test(is_admin_or_manager, login_url='login'), name='dispatch')
class AssetBulkImportView(View):
    template_name = 'assets/asset_bulk_import.html'

    def get(self, request):
        # Step 1: Select category, download template
        categories = AssetCategory.objects.all()
        selected_category = request.GET.get('category')
        step = request.GET.get('step', '1')
        context = {'categories': categories, 'selected_category': selected_category, 'step': step}
        return render(request, self.template_name, context)

    def post(self, request):
        # Step 2: Upload and preview file
        categories = AssetCategory.objects.all()
        selected_category = request.POST.get('category')
        step = request.POST.get('step', '2')
        file = request.FILES.get('import_file')
        preview_data = []
        errors = []
        columns = []
        import_file = request.POST.get('import_file')
        if step == '3':
            # Final confirmation: import assets
            # Re-read the file from temp storage
            file_path = default_storage.path(import_file)
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(columns, row))
                    preview_data.append(row_data)
            except Exception as e:
                errors.append(f'Failed to parse file: {e}')
            dynamic_fields = AssetCategoryField.objects.filter(category_id=selected_category)
            success_count = 0
            fail_count = 0
            fail_rows = []
            with transaction.atomic():
                for i, row in enumerate(preview_data):
                    try:
                        # Validate required fields
                        for field in dynamic_fields:
                            if field.required and not row.get(field.key):
                                raise ValueError(f'Missing required field {field.label}')
                        asset = Asset(
                            category_id=selected_category,
                            status=row.get('status', 'active'),
                            description=row.get('description', ''),
                        )
                        # Assign user if provided
                        assigned_to = row.get('assigned_to')
                        if assigned_to:
                            from users.models import User
                            user_obj = User.objects.filter(username=assigned_to).first()
                            if user_obj:
                                asset.assigned_to = user_obj
                                # Log assignment/transfer
                                log_audit(request.user, ASSIGN_ACTION, asset, f'Asset assigned to {user_obj.username} via bulk import (row {i+2})', related_user=user_obj)
                        # Dynamic fields
                        dyn_data = {}
                        for field in dynamic_fields:
                            dyn_data[field.key] = row.get(field.key)
                        asset.dynamic_data = dyn_data
                        asset.save()
                        success_count += 1
                        log_audit(request.user, 'create', asset, f'Asset imported via bulk import (row {i+2})')
                    except Exception as e:
                        fail_count += 1
                        fail_rows.append({'row': i+2, 'error': str(e)})
                # Audit log
                from assets.models import ExportLog
                ExportLog.objects.create(
                    user=request.user,
                    format='import',
                    columns=columns,
                    filters={'category': selected_category},
                    success=(fail_count == 0),
                    error_message='; '.join([f"Row {r['row']}: {r['error']}" for r in fail_rows]) if fail_rows else ''
                )
            context = {
                'categories': categories,
                'selected_category': selected_category,
                'step': 'done',
                'success_count': success_count,
                'fail_count': fail_count,
                'fail_rows': fail_rows,
            }
            return render(request, self.template_name, context)
        if not file or not selected_category:
            messages.error(request, 'Please select a category and upload a file.')
            return render(request, self.template_name, {'categories': categories, 'selected_category': selected_category, 'step': '1'})
        # Save file temporarily
        tmp_path = default_storage.save('tmp/' + file.name, file)
        file_path = default_storage.path(tmp_path)
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(columns, row))
                preview_data.append(row_data)
        except Exception as e:
            errors.append(f'Failed to parse file: {e}')
        # Validate rows (basic, more in confirm step)
        dynamic_fields = AssetCategoryField.objects.filter(category_id=selected_category)
        for i, row in enumerate(preview_data):
            for field in dynamic_fields:
                if field.required and not row.get(field.key):
                    errors.append(f'Row {i+2}: Missing required field {field.label}')
        context = {
            'categories': categories,
            'selected_category': selected_category,
            'step': step,
            'preview_data': preview_data,
            'columns': columns,
            'errors': errors,
            'import_file': tmp_path,
        }
        return render(request, self.template_name, context)

    def put(self, request):
        # Step 3: Confirm import (AJAX or form submit)
        # Not implemented yet
        pass

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'assets/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context['role'] = getattr(user, 'role', 'user')
        return context

@login_required
@require_GET
def dashboard_summary_api(request):
    from datetime import timedelta
    from django.utils import timezone
    user = request.user
    role = getattr(user, 'role', 'user')
    qs = Asset.objects.all()
    if role == 'user':
        qs = qs.filter(assigned_to=user)
    # KPIs
    total_assets = qs.count()
    active_assets = qs.filter(status='active').count()
    maintenance_assets = qs.filter(status='maintenance').count()
    retired_assets = qs.filter(status='retired').count()
    lost_assets = qs.filter(status='lost').count()
    assigned_assets = qs.filter(assigned_to__isnull=False).count()
    unassigned_assets = qs.filter(assigned_to__isnull=True).count()
    # Warranty expiry (assuming dynamic_data has 'warranty_expiry' as ISO date string)
    soon = timezone.now() + timedelta(days=30)
    warranty_expiring_soon = qs.filter(dynamic_data__warranty_expiry__lte=soon.date().isoformat(), dynamic_data__warranty_expiry__gte=timezone.now().date().isoformat()).count()
    # Transferred assets (assuming status or audit log, fallback to status='transferred' if exists)
    transferred_assets = qs.filter(status='transferred').count() if 'transferred' in dict(Asset.STATUS_CHOICES) else 0
    # By category
    by_category = {}
    for cat in AssetCategory.objects.all():
        cat_qs = qs.filter(category=cat)
        by_category[cat.name] = cat_qs.count()
    # Trends (example: monthly change in total assets)
    month_ago = timezone.now() - timedelta(days=30)
    total_assets_month_ago = qs.filter(created_at__lte=month_ago).count()
    total_assets_monthly_change = None
    if total_assets_month_ago:
        total_assets_monthly_change = f"{((total_assets - total_assets_month_ago) / total_assets_month_ago) * 100:.1f}%"
    else:
        total_assets_monthly_change = "N/A"
    # TODO: Consider logging dashboard summary API access for auditability
    return JsonResponse({
        'kpis': {
            'total_assets': total_assets,
            'active_assets': active_assets,
            'maintenance_assets': maintenance_assets,
            'retired_assets': retired_assets,
            'lost_assets': lost_assets,
            'assigned_assets': assigned_assets,
            'unassigned_assets': unassigned_assets,
            'warranty_expiring_soon': warranty_expiring_soon,
            'transferred_assets': transferred_assets,
        },
        'by_category': by_category,
        'trends': {
            'total_assets_monthly_change': total_assets_monthly_change,
        },
        'role': role,
        'user_id': user.id,
    })

@login_required
@require_GET
def dashboard_activity_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.all().order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    elif role == 'manager':
        # Example: managers see team logs if available
        pass
    logs = logs[:20]
    data = [
        {
            'user': str(log.user) if log.user else '',
            'action': log.action,
            'asset': str(log.asset) if log.asset else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in logs
    ]
    # TODO: Consider logging dashboard activity API access for auditability
    return JsonResponse({'activity': data})

@login_required
@require_GET
def dashboard_scan_logs_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='view').order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    elif role == 'manager':
        # Example: managers see team logs if available
        pass
    logs = logs[:5]
    data = [
        {
            'user': str(log.user) if log.user else '',
            'asset': str(log.asset) if log.asset else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
        }
        for log in logs
    ]
    # TODO: Consider logging dashboard scan logs API access for auditability
    return JsonResponse({'scan_logs': data})

@login_required
@require_GET
def dashboard_chart_data_api(request):
    from django.db.models import Count
    from django.utils import timezone
    import calendar
    import datetime
    user = request.user
    role = getattr(user, 'role', 'user')
    qs = Asset.objects.all()
    if role == 'user':
        qs = qs.filter(assigned_to=user)
    chart = request.GET.get('chart')
    if not chart or chart not in {'category', 'acquisition', 'department', 'location', 'depreciation'}:
        return JsonResponse({'error': 'Invalid or missing chart type'}, status=400)
    data = []
    labels = []
    # 1. Assets by Category
    if chart == 'category':
        agg = qs.values('category__name').annotate(count=Count('id')).order_by('-count')
        labels = [a['category__name'] for a in agg]
        data = [a['count'] for a in agg]
        return JsonResponse({'chart': 'assets_by_category', 'labels': labels, 'data': data, 'role': role})
    # 2. Asset Acquisition Over Time (last 12 months)
    elif chart == 'acquisition':
        now = timezone.now()
        months = [(now - datetime.timedelta(days=30*i)).strftime('%Y-%m') for i in reversed(range(12))]
        month_counts = {m: 0 for m in months}
        for asset in qs:
            m = asset.created_at.strftime('%Y-%m')
            if m in month_counts:
                month_counts[m] += 1
        labels = list(month_counts.keys())
        data = list(month_counts.values())
        return JsonResponse({'chart': 'acquisition_over_time', 'labels': labels, 'data': data, 'role': role})
    # 3. Assets by Department (dynamic field)
    elif chart == 'department':
        from django.db.models import Count, Value as V
        from django.db.models.functions import Coalesce
        from django.db.models.expressions import RawSQL
        try:
            # Use KeyTextTransform for JSONField key extraction (Django >=3.1)
            from django.db.models.functions import Cast
            from django.db.models import CharField
            qs_with_dept = qs.annotate(
                department=Cast('dynamic_data__department', CharField())
            )
            agg = qs_with_dept.values('department').annotate(count=Count('id')).order_by('-count')
            labels = [a['department'] or 'Unspecified' for a in agg]
            data = [a['count'] for a in agg]
        except Exception:
            # Fallback to Python loop if ORM fails
            dept_counts = {}
            for asset in qs:
                dept = asset.dynamic_data.get('department', 'Unspecified')
                dept_counts[dept] = dept_counts.get(dept, 0) + 1
            labels = list(dept_counts.keys())
            data = list(dept_counts.values())
        return JsonResponse({'chart': 'assets_by_department', 'labels': labels, 'data': data, 'role': role})
    # 4. Assets by Location (dynamic field)
    elif chart == 'location':
        from django.db.models import Count, Value as V
        from django.db.models.functions import Coalesce
        from django.db.models.expressions import RawSQL
        try:
            from django.db.models.functions import Cast
            from django.db.models import CharField
            qs_with_loc = qs.annotate(
                location=Cast('dynamic_data__location', CharField())
            )
            agg = qs_with_loc.values('location').annotate(count=Count('id')).order_by('-count')
            labels = [a['location'] or 'Unspecified' for a in agg]
            data = [a['count'] for a in agg]
        except Exception:
            loc_counts = {}
            for asset in qs:
                loc = asset.dynamic_data.get('location', 'Unspecified')
                loc_counts[loc] = loc_counts.get(loc, 0) + 1
            labels = list(loc_counts.keys())
            data = list(loc_counts.values())
        return JsonResponse({'chart': 'assets_by_location', 'labels': labels, 'data': data, 'role': role})
    # 5. Depreciation/Value Trend (robust, using explicit depreciation fields)
    elif chart == 'depreciation':
        now = timezone.now()
        months = [(now - datetime.timedelta(days=30*i)).replace(day=1).strftime('%Y-%m') for i in reversed(range(12))]
        month_values = {m: 0 for m in months}
        has_value_data = False
        for asset in qs:
            # Only include assets with all required depreciation fields
            if not (asset.purchase_value and asset.purchase_date and asset.useful_life_years):
                continue
            if asset.depreciation_method != 'straight_line':
                continue  # Only support straight-line for now
            purchase_value = float(asset.purchase_value)
            purchase_date = asset.purchase_date
            useful_life_months = asset.useful_life_years * 12
            monthly_depreciation = purchase_value / useful_life_months
            for m in months:
                year, month = map(int, m.split('-'))
                period_start = datetime.datetime(year, month, 1, tzinfo=now.tzinfo)
                if purchase_date > period_start.date():
                    continue  # Asset not yet acquired
                months_elapsed = (period_start.year - purchase_date.year) * 12 + (period_start.month - purchase_date.month)
                if months_elapsed < 0:
                    continue
                depreciated_value = max(purchase_value - monthly_depreciation * months_elapsed, 0)
                # Asset cannot depreciate below zero
                month_values[m] += depreciated_value
                has_value_data = True
        labels = list(month_values.keys())
        data = list(month_values.values())
        if not has_value_data:
            return JsonResponse({'chart': 'depreciation_trend', 'labels': labels, 'data': [], 'role': role, 'message': 'No depreciable asset data available for trend.'})
        return JsonResponse({'chart': 'depreciation_trend', 'labels': labels, 'data': data, 'role': role})
    # Should not reach here due to earlier validation
    return JsonResponse({'error': 'Invalid chart type'}, status=400)

def paginate_logs(logs, request, default_size=10, max_size=50):
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', default_size))
        page_size = min(max(page_size, 1), max_size)
    except Exception:
        page = 1
        page_size = default_size
    paginator = Paginator(logs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return page_obj, paginator

@login_required
@require_GET
def recent_added_assets_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action__in=['create', 'add']).order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_added_assets': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_scans_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='scan').order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_scans': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_transfers_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='assign').order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'from_user': str(log.user) if log.user else '',
            'to_user': str(log.related_user) if log.related_user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_transfers': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def recent_maintenance_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.filter(action='maintenance').order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    page_obj, paginator = paginate_logs(logs, request)
    data = [
        {
            'asset_id': log.asset.id if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        }
        for log in page_obj
    ]
    return JsonResponse({'recent_maintenance': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
@require_GET
def full_audit_log_api(request):
    user = request.user
    role = getattr(user, 'role', 'user')
    logs = AuditLog.objects.all().order_by('-timestamp')
    if role == 'user':
        logs = logs.filter(user=user)
    elif role == 'manager':
        pass
    action = request.GET.get('action')
    if action:
        logs = logs.filter(action=action)
    asset_id = request.GET.get('asset_id')
    if asset_id:
        logs = logs.filter(asset__id=asset_id)
    user_id = request.GET.get('user_id')
    if user_id:
        logs = logs.filter(user__id=user_id)
    page_obj, paginator = paginate_logs(logs, request, default_size=20, max_size=100)
    data = [
        {
            'id': log.id,
            'action': log.action,
            'asset_id': log.asset.id if log.asset else None,
            'asset_name': str(log.asset) if log.asset else '',
            'user': str(log.user) if log.user else '',
            'related_user': str(log.related_user) if log.related_user else '',
            'timestamp': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
            'metadata': log.metadata,
        }
        for log in page_obj
    ]
    return JsonResponse({'audit_log': data, 'page': page_obj.number, 'num_pages': paginator.num_pages, 'total': paginator.count})

@login_required
def user_assets_api(request):
    user = request.user
    assets = Asset.objects.filter(assigned_to=user)
    data = []
    for asset in assets:
        data.append({
            'name': str(asset),
            'serial': asset.dynamic_data.get('serial_number', ''),
            'assigned': str(asset.assigned_to) if asset.assigned_to else '',
            'status': asset.status,
        })
    return JsonResponse({'assets': data})

@login_required
def user_activity_api(request):
    user = request.user
    logs = AuditLog.objects.filter(user=user).order_by('-timestamp')[:20]
    data = []
    for log in logs:
        data.append({
            'action': log.action,
            'asset': str(log.asset) if log.asset else '',
            'time': localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
            'details': log.details,
        })
    return JsonResponse({'logs': data})
